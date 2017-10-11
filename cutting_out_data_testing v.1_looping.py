import openpyxl
import datetime
from openpyxl.chart import ScatterChart, Reference, Series


def milestone_swimlane(col, start_row, project_number, newwb):
       
        newsheet = newwb.active

        wb = openpyxl.load_workbook('C:\\Users\\Standalone\\Documents\\Old portfolio masters\\compiled_master_2017-04-18_Q4 Jan – Mar 2017.xlsx')
        sheet = wb.active
        x = 1
        for i in range (88, 268, 6):
                val = sheet.cell(row=i, column=col).value
                newsheet.cell(row=x, column=1, value=val)
                x += 1

        x = 1
        for i in range (89, 269, 6):
                val = sheet.cell(row=i, column=col).value
                newsheet.cell(row=x, column=2, value=val)
                x += 1

        today = datetime.datetime.today()
        for i in range (start_row, start_row + 31):
                time_line_date = newsheet.cell(row=i, column=2).value
                try:
                        difference = (time_line_date - today).days
                        print(difference)
                        if difference in range (1, 5000):
                                newsheet.cell(row=i, column=3, value=(time_line_date - today))
                except TypeError:
                        pass
                
        for i in range (start_row, start_row + 31):
                newsheet.cell(row=i, column=4, value=project_number)

        chart = ScatterChart()
        chart.title = "Scatter Chart"
        chart.style = 1 
        chart.x_axis.title = 'Date'
        chart.y_axis.title = 'Project No'

        xvalues = Reference(newsheet, min_col=3, min_row=1, max_row=30)
        for i in range(1, 31):
            values = Reference(newsheet, min_col=4, min_row=1, max_row=30)
            series = Series(values, xvalues, title_from_data=True)
            chart.series.append(series)

        newsheet.add_chart(chart, "E10")
        
        return newwb
       

# open new workbook
# open source workbook
st_row = 1
proj_num = 1
newwb = openpyxl.Workbook()
for column in range(2, 33):
        wb = milestone_swimlane(column, st_row, proj_num, newwb)
        proj_num = 1
        colum
wb.save('C:\\Users\\Standalone\\Documents\\cutting_data_test2.xlsx')



