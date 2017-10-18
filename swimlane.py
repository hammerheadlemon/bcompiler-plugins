import os
import openpyxl
import datetime
from openpyxl.chart import ScatterChart, Reference, Series

HOME = os.path.abspath(os.path.expanduser('~'))
DESKTOP = os.path.join(HOME, 'Desktop')

NUMBER_OF_PROJECTS = 20


def milestone_swimlane(start_row, project_number, newwb, block_start_row=90,
                       interested_range=365):
    newsheet = newwb.active
    col = project_number + 1
    start_row = start_row + 1

    wb = openpyxl.load_workbook(
        os.path.join(
            DESKTOP, 'Q2_1718_master.xlsx'))
    sheet = wb.active

    # print project title
    newsheet.cell(row=start_row - 1, column=1, value=sheet.cell(row=1, column=col).value)
    print(sheet.cell(row=1, column=col).value)

    x = start_row
    for i in range(block_start_row, 269, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=1, value=val)
        x += 1
    x = start_row
    for i in range(block_start_row + 1, 270, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=2, value=val)
        x += 1

    today = datetime.datetime.today()
    current_row = start_row
    for i in range(91, 269, 6):
        time_line_date = sheet.cell(row=i, column=col).value
        try:
            difference = (time_line_date - today).days
            if difference in range(1, interested_range):
                newsheet.cell(row=current_row, column=3, value=difference)
        except TypeError:
            pass
        finally:
            current_row += 1

    for i in range(start_row, start_row + 30):
        newsheet.cell(row=i, column=4, value=project_number)

    return newwb, start_row


def _segment_series():
    cut = dict(
        sobc=1,
        obc=1,
        ds1=4,
        fbc=1,
        ds2=4,
        ds3=4,
        free=8
    )
    for item in cut.items():
        yield item

def _series_producer(sheet, start_row, step):
    xvalues = Reference(sheet, min_col=3, min_row=start_row, max_row=start_row + step)
    values = Reference(sheet, min_col=4, min_row=start_row, max_row=start_row + step)
    series = Series(values, xvalues)
    new_start = start_row + step + 1
    return series, new_start


def _start_cells():
    total_rows = NUMBER_OF_PROJECTS * 30
    for i in range(2, total_rows, 31):
        yield i


def _row_calc_chart(start_row):
    return start_row + 29


def _row_calc(project_number):
    if project_number == 1:
        return 1, 1
    if project_number == 2:
        return 2, 32
    else:
        return (project_number, (project_number + 30) + ((project_number - 2) * 30))


def main():
    wb = openpyxl.Workbook()
    start_generator = _start_cells()
    segment_series_generator = _segment_series()
    for p in range(1, 31):
        proj_num, st_row = _row_calc(p)
        wb = milestone_swimlane(st_row, proj_num, wb, block_start_row=90, interested_range=365)[0]

    chart = ScatterChart()
    chart.title = "Swimlane Chart"
    chart.style = 1
    chart.x_axis.title = 'Days from Today'
    chart.y_axis.title = 'Project No'

    derived_end = 2

    for p in range(1, NUMBER_OF_PROJECTS):
        for i in range(1, 8):  # 8 here is hard-coded number of segments within a project series (ref: dict in _segment_series()
            if i == 1:
                inner_start_row = derived_end
            else:
                inner_start_row = derived_end
            _inner_step = next(segment_series_generator)[1]
            series, derived_end = _series_producer(wb.active, inner_start_row, _inner_step)
            if _inner_step == 1:
                series.marker.symbol = "triangle"
                series.marker.graphicalProperties.solidFill = "01a852"
            else:
                series.marker.symbol = "square"
                series.marker.graphicalProperties.solidFill = "FF0000"
            series.marker.size = 10
            chart.series.append(series)
        start_generator = _start_cells()
        segment_series_generator = _segment_series()
        derived_end = derived_end + 1

    wb.active.add_chart(chart, "E1")
    wb.save(os.path.join(DESKTOP, 'output.xlsx'))


if __name__ == "__main__":
    main()
