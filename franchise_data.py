from openpyxl import load_workbook, Workbook

q1 = load_workbook('source_files/compiled_master_2017-10-11_Q1 Franchising.xlsx')
q2 = load_workbook('source_files/compiled_master_2017-10-11_Q2 Franchising.xlsx')

ws1 = q1.active
ws2 = q2.active

output_wb = Workbook()
ws_output = output_wb.active

for c in range(2, ws1.max_column):
    ws_output.cell(row=1, column=c, value=ws1.cell(row=1, column=c).value)


def single_project(col_letter):
    cells = ws1[col_letter + '280':col_letter + '794']
    for c in enumerate(cells, start=2):
        ws_output.cell(row=c[0], column=c[1][0].col_idx, value=c[1][0].value)


single_project('A')

for letter in 'BCDEFG':
    single_project(letter)

output_wb.save('test.xlsx')
