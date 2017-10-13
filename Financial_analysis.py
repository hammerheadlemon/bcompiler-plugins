import openpyxl # module required to work with Excel spreadsheets
import datetime # module for working with dates
import os
from openpyxl.chart import ScatterChart, Reference, Series # module for creating charts in Excel

newwb = openpyxl.Workbook() # opens a new Excel workbook
newsheet = newwb.active # asks for all information to be passed to first Excel sheet

os.chdir('source_files')

wb1 = openpyxl.load_workbook('compiled_master_2017-10-11_Q1 Franchising.xlsx') # stores Q1 master spreadsheet as a variable called wb1
wb2 = openpyxl.load_workbook('compiled_master_2017-10-11_Q2 Franchising.xlsx') # store Q2 master spreadhseet as a variable called wb2

sheet1 = wb1.active # specifies the first worksheet within the Q1 master spreadsheet.
sheet2 = wb2.active # specifies the first worksheet within the Q2 master spreadsheet.

# set the key in column A
finance_info_key1 = sheet1.cell(row=1, column=1).value
newsheet.cell(row=1, column=1, value=finance_info_key1)

x = 2
for i in range (280, 794):
    finance_info_key2 = sheet1.cell(row=i, column=1).value
    newsheet.cell(row=x, column=1, value=finance_info_key2)
    x += 1

# puts project(s) names in first row from column B onwards

#x = 2
#for i in range(2, 8):
#    val = sheet1.cell(row=1, column=i).value
#    newsheet.cell(row=x, column=2, value=val)
#    x += 1

# put project name in first row of column be
project_name = sheet1.cell(row=1, column = 2).value
newsheet.cell(row=1, column=2, value=project_name)

# iterates through Q1 master taking project finance values. placing them in column B
x = 2
for i in range (280, 794):
    Q1_financials = sheet1.cell(row=i, column=2).value
    newsheet.cell(row=x, column=2, value=Q1_financials)
    x += 1

project_name2 = sheet2.cell(row=1, column = 2).value
newsheet.cell(row=1, column=3, value=project_name2)

# iterates through Q2 master taking project finance values. placing them in column c
x = 2
for i in range (280, 794):
    Q2_financials = sheet2.cell(row=i, column=2).value
    newsheet.cell(row=x, column=3, value=Q2_financials)
    x += 1

# iterate through and calculate difference between quarter returns

for i in range (19, 515):
    try:
        Q1_data = (newsheet.cell(row=i, column=2).value)
        Q2_data = (newsheet.cell(row=i, column=3).value)
        financial_difference = Q1_data - Q2_data
        newsheet.cell(row=i, column=4, value=financial_difference)
    except TypeError:
        pass

'''
time_line_date = newsheet.cell(row=i, column=2).value
                try:
                        difference = (time_line_date - lst_quarter).days
                        print(difference)
                        if difference in range (1, 365):
                                newsheet.cell(row=i, column=3, value=(time_line_date - lst_quarter))
                except TypeError:
                        pass

#newsheet.cell(row=19, column=2).value
#newsheet.cell(row=19, column=3).value

        # TODO record different milestones as different variables. e.g 'SOBC', 'OBC', 'FBC' and rows 17 - 20 as 'project_milestones'

        # iterates through master taking 'forecast/actual dates' for MM values [note - change from first programme written on Monday 9 Oct, which was picking up
        # 'original/baseline' dates. Placing them in column B
        x = 1
        for i in range (91, 271, 6):
                val = sheet.cell(row=i, column=col).value
                newsheet.cell(row=x, column=2, value=val)
                x += 1

        # calculates the time delta between milestone dates and today. The only prints in worksheet those timedeltas in the next 365 days i.e. year period. [note
        # calculations can be altered to suit the start date and time period required]. placing them in column C.
        lst_quarter = datetime.datetime(2017,6,30,0,0)
        for i in range (1, 31):
                time_line_date = newsheet.cell(row=i, column=2).value


        # prints the value 1 in column D. required for the excel scatter chart.
        for i in range (1, 31):
                newsheet.cell(row=i, column=4, value=1)

        # TODO record different milestones as different variables. e.g 'SOBC', 'OBC', 'FBC' and rows 17 - 20 as 'project_milestones'. So that
        # they can be displayed on chart as different symbols

                #SOBC = newsheet.cell(row=1, column=1).value - this worked, but didn't like a range.

        #builds the scatterchart
        chart = ScatterChart()
        chart.title = "Project Planning"
        chart.x_axis.title = 'Date'
        chart.y_axis.title = 'Project No'

        xvalues = Reference(newsheet, min_col=3, min_row=1, max_row=30)
        values = Reference(newsheet, min_col=4, min_row=1, max_row=30)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)

        s1 = chart.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

        newsheet.add_chart(chart, "G8")
'''

# writes/saves all information to document as specified in the string
newwb.save('finance_testing.xlsx')


#milestone_swimlane(2)
