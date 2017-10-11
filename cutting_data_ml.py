import openpyxl
import datetime
from openpyxl.chart import ScatterChart, Reference, Series


def milestone_swimlane(start_row, project_number, newwb):

    newsheet = newwb.active
    col = project_number + 1
    start_row = start_row + 1

    wb = openpyxl.load_workbook('/home/lemon/Downloads/compiled_master_2017-07-18_Q1 Apr - Jun 2017 FOR Q2 COMMISSION DO NOT CHANGE.xlsx')
    sheet = wb.active

    # print project title
    newsheet.cell(row=start_row - 1, column=1, value=sheet.cell(row=1, column=col).value)
    print(sheet.cell(row=1, column=col).value)

    x = start_row
    for i in range(90, 269, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=1, value=val)
        x += 1
    x = start_row
    for i in range(91, 269, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=2, value=val)
        x += 1

    today = datetime.datetime.today()
    current_row = start_row
    for i in range(91, 269, 6):
        time_line_date = sheet.cell(row=i, column=col).value
        try:
            difference = (time_line_date - today).days
            if difference in range(1, 5000):
                newsheet.cell(row=current_row, column=3, value=difference)
        except TypeError:
                pass
        finally:
            current_row += 1

    for i in range(start_row, start_row + 30):
        newsheet.cell(row=i, column=4, value=project_number)


#   chart = ScatterChart()
#   chart.title = "Scatter Chart"
#   chart.style = 1
#   chart.x_axis.title = 'Date'
#   chart.y_axis.title = 'Project No'
#
#   xvalues = Reference(newsheet, min_col=3, min_row=1, max_row=30)
#   for i in range(1, 31):
#       values = Reference(newsheet, min_col=4, min_row=1, max_row=30)
#       series = Series(values, xvalues, title_from_data=True)
#       chart.series.append(series)
#
#   newsheet.add_chart(chart, "A10")

    return newwb


def _row_calc(project_number):
    if project_number == 1:
        return (1, 1)
    if project_number == 2:
        return (2, 32)
    else:
        return (project_number, (project_number + 30) + ((project_number - 2) * 30))


def main():
    wb = openpyxl.Workbook()
    for p in range(1, 31):
        proj_num, st_row = _row_calc(p)
        wb = milestone_swimlane(st_row, proj_num, wb)
    wb.save('output.xlsx')


if __name__ == "__main__":
    main()


